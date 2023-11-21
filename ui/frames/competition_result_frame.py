import os
import sys
import tempfile
from tkinter import ttk
from datetime import datetime
import openpyxl

from fpdf import FPDF
from data.competition import SORTING_FUNCTION
from data.match import Match
from idlelib.tooltip import Hovertip
from .default_frame import DefaultFrame


class CompetitionResultFrame(DefaultFrame):
    def __init__(self, container, parent):
        super().__init__(container, parent)

        self.label_list = []
        # field options
        options = {"padx": 5, "pady": 0}
        # user input
        self.generate_button = ttk.Button(
            self, text="PDF erzeugen", command=self.actionPrint
        )
        Hovertip(
            self.generate_button,
            "Erzeugt einen druckbaren Bericht",
        )
        self.close_button = ttk.Button(self, text="Beenden", command=self.actionClose)
        Hovertip(
            self.close_button,
            "Wettbewerb beenden",
        )
        self.columnconfigure(3, weight=1)
        self.grid(column=1, row=0, padx=5, pady=5, sticky="nsew")

    def actionPrint(self):
        if self.league:
            self.actionXLS()
        else:
            self.actionPDF()
        # self.remove_current_competition()

    def actionClose(self):
        self.remove_current_competition()
        self.proceed()

    def actionXLS(self):
        self._makeLeagueXLS()
        # self.generate_button["state"] = "disabled"

    def actionPDF(self):
        self._makeCompetionPDF()
        # self.generate_button["state"] = "disabled"

    def remove_current_competition(self):
        self.competition.active = False
        self.competitions.save()
        self.active_competitions.remove(self.competition)
        self.competition = None

    def discard_current_competition(self):
        self.competitions.remove(self.competition.id)
        if self.competition.league:
            self.leagues.remove(self.competition.league)
            self.leagues.save()
        self.remove_current_competition()

    def get_sorted_results(self) -> list[Match]:
        return sorted(
            [self.matches[key] for key in self.competition.entries],
            key=SORTING_FUNCTION[self.competition.modus]["key"],
            reverse=SORTING_FUNCTION[self.competition.modus]["reverse"],
        )

    def _makeLeagueXLS(self):
        FILENAME = "extern/formular_ligaergebnisse_kr_bez_ll.xlsx"
        LANDESLIGA_GRUPPE = "Z2"
        BEZIRKSLIGA_GRUPPE = "Z3"
        KREISLIGA_GRUPPE = "Z4"
        DISZIPLIN = "E5"
        DATUM = "R5"
        HEIM = "B8"
        GAST = "O8"
        SCHUSSZAHL = "H10"
        if not os.path.exists(FILENAME):
            raise Exception(msg="Formular nicht gefunden")
        workbook = openpyxl.load_workbook(filename=FILENAME)

        sheet = workbook.active
        sheet[DATUM] = self.competition.date
        sheet[DISZIPLIN] = self.league.discipline
        sheet[
            HEIM
        ] = f"{self.clubs[self.league.home_club].name} - {self.teams[self.league.home_team].name}"
        sheet[
            GAST
        ] = f"{self.clubs[self.league.guest_club].name} - {self.teams[self.league.guest_team].name}"
        sheet[SCHUSSZAHL] = self.competition.count
        if self.league.league == "Kreisliga":
            sheet[KREISLIGA_GRUPPE] = self.league.group
        elif self.league.league == "Bezirksliga":
            sheet[BEZIRKSLIGA_GRUPPE] = self.league.group
        elif self.league.league == "Landesliga":
            sheet[LANDESLIGA_GRUPPE] = self.league.group
        row_shooter_1 = 12
        column_name_home = 4
        column_name_guest = 18
        column_last_series_home = 12
        column_first_serie_home = column_last_series_home - self.competition.count // 10
        column_last_series_guest = 26
        column_first_serie_guest = (
            column_last_series_guest - self.competition.count // 10
        )
        home_entries = filter(
            lambda x: x.club == self.league.home_club,
            [self.matches[key] for key in self.competition.entries],
        )
        guest_entries = filter(
            lambda x: x.club == self.league.guest_club,
            [self.matches[key] for key in self.competition.entries],
        )
        for n, entry in enumerate([entry for entry in home_entries]):
            sheet.cell(
                row=row_shooter_1 + n, column=column_name_home
            ).value = entry.shooter.name
            for m, series in enumerate(entry.series):
                sheet.cell(
                    row=row_shooter_1 + n, column=column_first_serie_home + m
                ).value = series.summe_ganz
        for n, entry in enumerate([entry for entry in guest_entries]):
            sheet.cell(
                row=row_shooter_1 + n, column=column_name_guest
            ).value = entry.shooter.name
            for m, series in enumerate(entry.series):
                sheet.cell(
                    row=row_shooter_1 + n, column=column_first_serie_guest + m
                ).value = series.summe_ganz

        filedate = datetime.strptime(self.competition.date, "%d.%m.%Y")
        filename = f"""{datetime.strftime(filedate, "%y%m%d")}_Liga"""
        testfilename = f"{filename}.xlsx"
        n = 1
        while os.path.exists(os.path.join("output", testfilename)):
            testfilename = f"{filename}_{n}.xlsx"
            n += 1
        workbook.save(filename=os.path.join("output", testfilename))
        if not sys.platform == "linux":
            os.startfile(os.path.join("output", testfilename))

    def _makeCompetionPDF(self):
        filedate = datetime.strptime(self.competition.date, "%d.%m.%Y")
        filepath = self.competition.name.replace(" ", "_")
        filename = f"""{datetime.strftime(filedate, "%y%m%d")}_{self.competition.type_of_target}"""
        outdir = os.getcwd()
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.text(
            10,
            20,
            f"Ergebnisliste {self.competition.name} {self.competition.date}",
        )
        pdf.set_font("Helvetica", "", 10)

        for n, entry in enumerate(self.get_sorted_results()):
            pdf.text(10, 30 + n * 5, str(n + 1) + ".")
            pdf.text(15, 30 + n * 5, entry.shooter.name)
            if self.competition.modus in ["Bestes Ergebnis", "Bestes Ergebnis Zehntel"]:
                for j, series in enumerate(entry.series):
                    pdf.text(
                        (190 - 120) + j * 10,
                        30 + n * 5,
                        f"{series.summe if self.competition.decimal else series.summe_ganz}",
                    )
            pdf.text(
                190,
                30 + n * 5,
                str(SORTING_FUNCTION[self.competition.modus]["key"](entry)),
            )
        newpath = os.path.join(outdir, "output", "competitions", filepath)
        if not os.path.exists(newpath):
            os.makedirs(newpath)
        testfilename = f"{filename}.pdf"
        n = 1
        while os.path.exists(os.path.join(newpath, testfilename)):
            testfilename = f"{filename}_{n}.pdf"
            n += 1
        pdf.output(os.path.join(newpath, testfilename), "F")

        if not sys.platform == "linux":
            os.startfile(os.path.join(newpath, testfilename))

    def reset(self):
        if not self.competition.entries:
            self.discard_current_competition()
            return self.proceed()
        self.generate_button["state"] = "normal"
        self.activate_ok_button()
        for label in self.label_list:
            label.destroy()
        n = 0
        m = 0
        if self.league:
            home_entries = filter(
                lambda x: x.club == self.league.home_club,
                [self.matches[key] for key in self.competition.entries],
            )
            guest_entries = filter(
                lambda x: x.club == self.league.guest_club,
                [self.matches[key] for key in self.competition.entries],
            )
            header_home = ttk.Label(self, text="Heim")
            header_home.grid(row=0, column=0, columnspan=3)
            self.label_list.append(header_home)
            header_guest = ttk.Label(self, text="Gast")
            header_guest.grid(row=0, column=4, columnspan=3)
            self.label_list.append(header_guest)
            for n, entry in enumerate([entry for entry in home_entries]):
                rank_label = ttk.Label(self, text=f"{n+1}")
                rank_label.grid(row=n + 1, column=0)
                self.label_list.append(rank_label)
                name_label = ttk.Label(self, text=f"{entry.shooter.name}")
                name_label.grid(row=n + 1, column=1)
                self.label_list.append(name_label)
                result_label = ttk.Label(
                    self,
                    text=f"{entry.summe_ganz}",
                )
                result_label.grid(row=n + 1, column=2)
                self.label_list.append(result_label)
            for m, entry in enumerate([entry for entry in guest_entries]):
                result_label = ttk.Label(
                    self,
                    text=f"{entry.summe_ganz}",
                )
                result_label.grid(row=m + 1, column=4)
                self.label_list.append(result_label)
                name_label = ttk.Label(self, text=f"{entry.shooter.name}")
                name_label.grid(row=m + 1, column=5)
                self.label_list.append(name_label)
                rank_label = ttk.Label(self, text=f"{m+1}")
                rank_label.grid(row=m + 1, column=6)
                self.label_list.append(rank_label)

            pass
        else:
            for n, entry in enumerate(self.get_sorted_results()):
                rank_label = ttk.Label(self, text=f"{n+1}")
                rank_label.grid(row=n, column=0)
                self.label_list.append(rank_label)
                name_label = ttk.Label(self, text=f"{entry.shooter.name}")
                name_label.grid(row=n, column=1)
                self.label_list.append(name_label)
                result_label = ttk.Label(
                    self,
                    text=f"{str(SORTING_FUNCTION[self.competition.modus]['key'](entry))}",
                )
                result_label.grid(row=n, column=2)
                self.label_list.append(result_label)
        self.generate_button.grid(
            row=max(n, m) + 2, column=0, columnspan=6, sticky="se"
        )
        self.close_button.grid(row=max(n, m) + 2, column=6, sticky="se")
